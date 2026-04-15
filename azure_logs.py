import requests
import base64
from azure.identity import ClientSecretCredential
from azure.monitor.query import LogsQueryClient
from datetime import datetime, timedelta, timezone
import pandas as pd
import os
from openpyxl.utils import get_column_letter

# ================================
# ----------- TENANT A (AZURE LOGS)
# ================================

resource_id = "/subscriptions/e55e5916-f6a5-42b6-b4e4-8e5de225fba0/resourceGroups/rg-nacha-csp-prod/providers/Microsoft.Insights/components/appi-nacha-csp-prod"

tenant_id_A = os.getenv("TENANT_ID_A")
client_id_A = os.getenv("CLIENT_ID_A")
client_secret_A = os.getenv("CLIENT_SECRET_A")

credential_A = ClientSecretCredential(
    tenant_id=tenant_id_A,
    client_id=client_id_A,
    client_secret=client_secret_A
)

client = LogsQueryClient(credential_A)

# ================================
# ----------- TIME RANGE
# ================================

end_utc = datetime.now(timezone.utc)
start_utc = end_utc - timedelta(hours=24)

start_utc_str = start_utc.strftime("%Y-%m-%dT%H:%M:%SZ")
end_utc_str = end_utc.strftime("%Y-%m-%dT%H:%M:%SZ")

print(f"\nUTC Window: {start_utc} → {end_utc}")

# ================================
# ----------- QUERIES
# ================================

query_1 = f"""
dependencies
| where timestamp between (datetime({start_utc_str}) .. datetime({end_utc_str}))
| where name == "POST /ACHCheckPrescreen/GetReport"
| where success != true
| project name, appId, target, success, resultCode, ["TimeStamp(UTC)"] = timestamp
| order by timestamp desc
"""

query_2 = f"""
dependencies
| where timestamp between (datetime({start_utc_str}) .. datetime({end_utc_str}))
| where name == "POST /ACHCheckPrescreen/GetReport"
| summarize 
    SuccessCount = countif(success == true),
    FailureCount = countif(success == false),
    TotalCount = count()
| extend SuccessRate = round((SuccessCount * 100.0) / TotalCount, 2)
"""

# ================================
# ----------- HELPERS
# ================================

def response_to_df(response):
    for table in response.tables:
        columns = [col.name if hasattr(col, "name") else col for col in table.columns]
        return pd.DataFrame(table.rows, columns=columns)
    return pd.DataFrame()

def remove_timezone(df):
    for col in df.columns:
        if pd.api.types.is_datetime64_any_dtype(df[col]):
            df[col] = df[col].apply(lambda x: x.replace(tzinfo=None) if pd.notnull(x) else x)
    return df

# ================================
# ----------- RUN QUERIES
# ================================

print("\nRunning Query 1...")
df1 = response_to_df(client.query_resource(resource_id, query_1, timespan=(start_utc, end_utc)))

print("\nRunning Query 2...")
df2 = response_to_df(client.query_resource(resource_id, query_2, timespan=(start_utc, end_utc)))

df1 = remove_timezone(df1)
df2 = remove_timezone(df2)

# ================================
# ----------- SAVE EXCEL (FIXED)
# ================================

folder_path = "Nacha_Daily_reports"
os.makedirs(folder_path, exist_ok=True)

today_date = datetime.now().strftime("%Y-%m-%d")
file_path = os.path.join(folder_path, f"Nacha-{today_date}.xlsx")

with pd.ExcelWriter(file_path, engine="openpyxl") as writer:
    df1.to_excel(writer, sheet_name="Failures", index=False)
    df2.to_excel(writer, sheet_name="Summary", index=False)

    workbook = writer.book

    # -------- FORMAT FAILURES SHEET --------
    sheet1 = writer.sheets["Failures"]

    for col_idx, col in enumerate(df1.columns, 1):
        max_length = max(
            df1[col].astype(str).map(len).max(),
            len(col)
        )
        sheet1.column_dimensions[get_column_letter(col_idx)].width = max_length + 2

    # Fix timestamp column format (Column F)
    for cell in sheet1["F"]:
        if cell.row != 1:
            cell.number_format = "yyyy-mm-dd hh:mm:ss"

    # -------- FORMAT SUMMARY SHEET --------
    sheet2 = writer.sheets["Summary"]

    for col_idx, col in enumerate(df2.columns, 1):
        max_length = max(
            df2[col].astype(str).map(len).max(),
            len(col)
        )
        sheet2.column_dimensions[get_column_letter(col_idx)].width = max_length + 2

print(f"\n Excel file generated: {file_path}")

# ================================
# ----------- TENANT B (GRAPH EMAIL)
# ================================

tenant_id_B = os.getenv("MAIL_TENANT_ID")
client_id_B = os.getenv("MAIL_CLIENT_ID")
client_secret_B = os.getenv("MAIL_CLIENT_SECRET")

sender_email = os.getenv("SENDER_EMAIL")
receiver_emails = os.getenv("RECEIVER_EMAILS")

email_list = [email.strip() for email in receiver_emails.split(",")]

to_recipients = [
    {"emailAddress": {"address": email}}
    for email in email_list
]

credential_B = ClientSecretCredential(
    tenant_id=tenant_id_B,
    client_id=client_id_B,
    client_secret=client_secret_B
)

token = credential_B.get_token("https://graph.microsoft.com/.default").token

# ================================
# ----------- EMAIL SEND
# ================================

url = f"https://graph.microsoft.com/v1.0/users/{sender_email}/sendMail"

with open(file_path, "rb") as f:
    file_content = base64.b64encode(f.read()).decode()

email_body = {
    "message": {
        "subject": f"NACHA Daily Report - {today_date}",
        "body": {
            "contentType": "Text",
            "content": f"""Hi Team,

Please find attached the NACHA daily report for {today_date} UTC.

Regards,
Sayan Karmakar"""
        },
        "toRecipients": to_recipients,
        "attachments": [
            {
                "@odata.type": "#microsoft.graph.fileAttachment",
                "name": os.path.basename(file_path),
                "contentBytes": file_content
            }
        ]
    }
}

headers = {
    "Authorization": f"Bearer {token}",
    "Content-Type": "application/json"
}

response = requests.post(url, headers=headers, json=email_body)

if response.status_code == 202:
    print(" Email sent !")
else:
    print(" Email failed:", response.status_code, response.text)
